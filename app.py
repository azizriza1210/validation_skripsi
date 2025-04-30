from flask import Flask, request, jsonify
from flask_cors import CORS
from groq import Groq
from save_chroma import save_to_chroma, rag
from prompt_design import ubah_prompt, buat_pertanyaan
from connect_mongo_db import show_data, login, simpan_data_user
from chat_to_sheet import save_to_sheet
from openpyxl import Workbook, load_workbook
import os
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi
from google import genai

username = None  
user_id = None  
collections_chatbot = None
db = None 

# Ini memastikan fungsi save_to_chroma() hanya dipanggil sekali
collection = save_to_chroma()
uri = "mongodb+srv://validasi:sociachat123@validasi.57paaa9.mongodb.net/?appName=Validasi"

# Create a new client and connect to the server
client = MongoClient(uri, server_api=ServerApi('1'))

# Send a ping to confirm a successful connection
try:
    client.admin.command('ping')
    print("--- Pinged your deployment. You successfully connected to MongoDB!  ---")
except Exception as e:
    print(e)

db = client['validation_database']
collections = db['validation_collection']


def create_app():
    # Inisialisasi Flask
    app = Flask(__name__)
    CORS(app)  # Agar bisa diakses dari frontend berbeda (seperteyi localhost:5173)

    # Inisialisasi Groq client
    # client_rag = Groq(api_key="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")  # Ganti "key" dengan API key Anda
    # client_rag = Groq(api_key="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")  # API saffa
    client_rag = genai.Client(api_key="AIzaSyA_98Dpic9pdz172faDcYnSpFsMAgLgQww") # mohammadazizriza

    # Model default GROQ
    GROQ_MODEL = "llama3-70b-8192"

    # Path file Excel
    EXCEL_FILE = "chat_log.xlsx"

    # Fungsi untuk menyimpan ke Excel
    def save_to_excel(prompt, response):
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Prompt", "Response"])  # Header
        
        ws.append([prompt, response])
        wb.save(EXCEL_FILE)

    @app.route("/chatbot/login", methods=["POST"])
    def login_sociachat():
        global username
        data = request.get_json(force=True)
        username = data.get("username")
        password = data.get("password")
        if username == "user1":
            nama = "Dr. Henike Primawanti, M.I.Pol."
        elif username == "user2":  
            nama = "Rino Adibowo, S.IP.,M.I.Po"
        elif username == "user3":
            nama = "Tatik Rohmawati, S.IP.,M.Si"
        else:
            nama = ""
        cek = login(username, password, nama)
        return cek
    
    @app.route("/chatbot/data")
    def get_data():
        global username, user_id, db, collections_chatbot
        collections_chatbot = collections
        result = show_data()
        tes = save_to_sheet(db, username)
        return "Selesai menyimpan data ke sheet"
    
    @app.route("/simpan-data", methods=["POST"])
    def simpan_data():
        data = request.get_json(force=True)
        hasil = simpan_data_user(data)
        return "hasil"

    @app.route("/chatbot/chat", methods=["POST"])
    def chatbot_chat():
        global user_id
        try:
            data = request.get_json(force=True)
            prompt = data.get("query")
            user_id = data.get("user_id")
            print(f"Prompt: {str(prompt)}")
            
            if not prompt:
                return jsonify({"error": "No query provided"}), 400
            
            # Ubah prompt menggunakan Groq
            optimal_prompt = ubah_prompt(str(prompt))
            # print(f"Optimal Prompt: {str(optimal_prompt)}")
            
            # Ambil data relevan dari RAG
            tweets = rag(collection, str(optimal_prompt))
            tweets_formatted = tweets if tweets else "Tidak ada informasi yang relevan ditemukan."
            # print(f"Tweets: {tweets_formatted}")

            # Buat prompt untuk LLM
            input_prompt = f"""INPUT Informasi yang tersedia: '{str(tweets_formatted)}' Pertanyaan: '{str(prompt)}' OUTPUT Harus Gunakan Bahasa Indonesia. ANSWER:"""
            
            # Stream response dan gabung jadi satu string
            # full_response = ""
            # for response in client_rag.chat.completions.create(
            #     model=GROQ_MODEL,
            #     messages=[{"role": "user", "content": input_prompt}],
            #     temperature=0.1,
            #     stream=True,
            # ):
            #     delta = response.choices[0].delta
            #     if delta.content:
            #         full_response += delta.content
            
            response = client_rag.models.generate_content(
                model="gemini-2.0-flash-lite",
                contents=input_prompt,
            )

            full_response = str(response.text)
            
            # Buat pertanyaan baru berdasarkan jawaban
            new_questions = buat_pertanyaan(prompt, full_response)
            
            new_data = {
                "chat": prompt,
                "response": full_response,
                "user_id": user_id
            }

            # Insert data ke koleksi
            insert_result = collections.insert_one(new_data)
            # Tampilkan ID dari dokumen yang baru saja dimasukkan
            print("Data berhasil dimasukkan dengan ID:", insert_result.inserted_id)
            
            return jsonify({"answer": full_response, "questions": list(new_questions)})
        
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    return app


# Jalankan Flask
if __name__ == "__main__":
    app = create_app()
    app.run(port=5000, debug=True)